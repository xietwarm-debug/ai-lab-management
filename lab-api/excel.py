# 需要先安装: pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from itertools import cycle

def generate_test_timetable(path="test_schedule_monday_friday.xlsx"):
    wb = Workbook()

    # =========================
    # Sheet1: 课表总览（周一~周五）
    # =========================
    ws = wb.active
    ws.title = "课表总览"

    weekdays = ["周一", "周二", "周三", "周四", "周五"]
    labs = ["D-C401", "D-C402", "D-C403", "D-C404", "D-C405", "D-C406"]
    periods = ["1-2节", "3-4节", "5-6节", "7-8节", "9-10节"]
    week_ranges = ["1-16周", "2-17周", "1-8周", "9-16周"]
    classes = ["计科23-1", "计科23-2", "软件23-1", "电信23-1", "自动化23-1", "人工智能23-1"]
    teachers = ["张老师", "李老师", "王老师", "赵老师", "刘老师", "陈老师"]
    courses = [
        "Python程序设计实验", "数据结构实验", "数据库系统实验", "计算机网络实验",
        "操作系统实验", "机器学习实验", "Web开发实验", "嵌入式系统实验"
    ]

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "实验室 \\ 星期"
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="D9E1F2")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for i, d in enumerate(weekdays, start=2):
        c = ws.cell(row=1, column=i, value=d)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.column_dimensions[c.column_letter].width = 34

    course_cycle = cycle(courses)
    teacher_cycle = cycle(teachers)
    class_cycle = cycle(classes)

    for r, lab in enumerate(labs, start=2):
        lab_cell = ws.cell(row=r, column=1, value=lab)
        lab_cell.font = Font(bold=True)
        lab_cell.fill = PatternFill("solid", fgColor="E2F0D9")
        lab_cell.alignment = Alignment(horizontal="center", vertical="center")
        lab_cell.border = border
        ws.row_dimensions[r].height = 110

        for c in range(2, 7):  # 周一~周五
            lines = []
            for p in periods:
                course = next(course_cycle)
                teacher = next(teacher_cycle)
                cls = next(class_cycle)
                week = week_ranges[(r + c) % len(week_ranges)]
                lines.append(f"{p} {course} | {week} | {teacher} | {cls}")
            txt = "\n".join(lines)
            cell = ws.cell(row=r, column=c, value=txt)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    # =========================
    # Sheet2: 课表明细（导入更友好）
    # =========================
    ws2 = wb.create_sheet("课表明细")
    headers = ["课程名称", "星期几", "节次", "周次", "实验室", "教师", "班级", "备注"]
    for i, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws2.column_dimensions[cell.column_letter].width = 18

    row = 2
    course_cycle = cycle(courses)
    teacher_cycle = cycle(teachers)
    class_cycle = cycle(classes)

    for lab in labs:
        for day in weekdays:
            for p in periods:
                ws2.cell(row=row, column=1, value=next(course_cycle))
                ws2.cell(row=row, column=2, value=day)
                ws2.cell(row=row, column=3, value=p)
                ws2.cell(row=row, column=4, value=week_ranges[row % len(week_ranges)])
                ws2.cell(row=row, column=5, value=lab)
                ws2.cell(row=row, column=6, value=next(teacher_cycle))
                ws2.cell(row=row, column=7, value=next(class_cycle))
                ws2.cell(row=row, column=8, value="测试数据")
                for col in range(1, 9):
                    ws2.cell(row=row, column=col).border = border
                row += 1

    wb.save(path)
    print(f"已生成: {path}")

if __name__ == "__main__":
    generate_test_timetable()
